import openpyxl
from openpyxl.cell import cell
import spacy
import json


print("Loading Spacy core")
nlp = spacy.load("en_core_web_lg")

search_term_file = "/Users/mattlavis/sites and projects/1. Online Tariff/process-hsen/_samples/search terms.xlsx"
processed_file = "/Users/mattlavis/sites and projects/1. Online Tariff/process-hsen/_samples/processed search terms.xlsx"
wb_obj = openpyxl.load_workbook(search_term_file)
sheet1 = wb_obj["Sheet1"]
max_row = sheet1.max_row
max_col = sheet1.max_column

facets = []
facets_dict = {}

sheet1.cell(row=1, column=4).value = "nouns"
sheet1.cell(row=1, column=5).value = "chunks"
sheet1.cell(row=1, column=6).value = "adjectives"

for i in range(2, max_row + 1):
    term = sheet1.cell(row=i, column=1).value
    terms = {}
    doc = nlp(term)
    terms["chunks"] = [chunk.text.lower() for chunk in doc.noun_chunks]
    terms["nouns"] = [token.lemma_.lower() for token in doc if token.pos_ == "NOUN"]
    terms["verbs"] = [token.lemma_.lower() for token in doc if token.pos_ == "VERB"]
    terms["adjectives"] = [token.lemma_.lower() for token in doc if token.pos_ == "ADJ"]

    nouns = " | ".join(token.lemma_.lower() for token in doc if token.pos_ == "NOUN")
    chunks = " | ".join(chunk.text.lower() for chunk in doc.noun_chunks)
    adjectives = " | ".join(token.lemma_.lower() for token in doc if token.pos_ == "ADJ")

    # Analyze syntax
    dependencies = []
    for token in doc:
        dependency = {
            token.text: [token.dep_]
        }
        dependencies.append(json.dumps(dependency))
    dependencies = " | ".join(dependencies)

    sheet1.cell(row=i, column=4).value = nouns
    sheet1.cell(row=i, column=5).value = chunks
    sheet1.cell(row=i, column=6).value = adjectives
    sheet1.cell(row=i, column=7).value = dependencies

    # if i > 50:
    #     break

    print(term)

wb_obj.save(processed_file)

# a = 1
